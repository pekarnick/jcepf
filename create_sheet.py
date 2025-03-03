import mysql.connector
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import Series
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from decimal import Decimal
from datetime import datetime
from dotenv import load_dotenv
import os
import sys

# Determinar el directorio base
if getattr(sys, 'frozen', False):  # Ejecutable
    BASE_DIR = os.path.dirname(sys.executable)
else:  # Script
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Cargar el archivo .env
env_path = os.path.join(BASE_DIR, '.env')
print(env_path)
if os.path.exists(env_path):
    load_dotenv(dotenv_path=env_path)
else:
    print(f"Archivo .env no encontrado en {env_path}")
    sys.exit(1)  # Salir si no hay .env

# Configuración de la conexión a MySQL
config = {
    'host': os.getenv('MYSQL_HOST'),
    'user': os.getenv('MYSQL_USER'),
    'password': os.getenv('MYSQL_PASSWORD'),
    'database': os.getenv('MYSQL_DATABASE'),
    'port': 3306,  # Añadir puerto explícito
    'raise_on_warnings': True  # Forzar reporte de errores
}

# Verificar que las variables estén cargadas
for key, value in config.items():
    if not value:
        print(f"Error: La variable de entorno {key.upper()} no está definida en .env")
        sys.exit(1)

# Función para ejecutar una consulta y devolver un DataFrame
def ejecutar_consulta(query):
    conn = None
    try:
        print(f"Intentando conectar con: host={config['host']}, user={config['user']}, db={config['database']}")
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        return pd.DataFrame(rows, columns=columns)
    except mysql.connector.Error as e:
        print(f"Error específico de MySQL: {e}")
        print(f"Código de error: {e.errno}, Mensaje: {e.msg}")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error inesperado: {type(e).__name__} - {str(e)}")
        return pd.DataFrame()
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

# Función para guardar DataFrame en una hoja de Excel
def guardar_en_excel(df, writer, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Obtener la hoja de trabajo
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Definir estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Side(border_style="thin", color="000000")

    # Aplicar formato a los encabezados
    for cell in worksheet[1]:  # Primera fila (encabezados)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Autoajustar el ancho de las columnas
    for col in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = col[0].column_letter  # Obtener letra de la columna
        worksheet.column_dimensions[col_letter].width = max_length + 5  # Aumentar de +2 a +5

    # Aplicar formato de moneda a columnas que contengan "Monto" o "Balance"
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if "Monto" in worksheet.cell(row=1, column=cell.column).value or "Balance" in worksheet.cell(row=1, column=cell.column).value:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Formato de moneda $

    # Agregar filtros automáticos a los encabezados
    last_column_letter = openpyxl.utils.get_column_letter(worksheet.max_column)
    filter_range = f"A1:{last_column_letter}1"  # Rango desde A1 hasta la última columna de la fila 1
    worksheet.auto_filter.ref = filter_range

    print(f"Formato y filtros aplicados a la hoja '{sheet_name}' correctamente.")

# Función para generar el resumen de ingresos, egresos y balance
def generar_resumen(df_ingresos, df_egresos, anio):
    # Asegurarse que los DataFrames tengan datos
    if df_ingresos.empty or df_egresos.empty:
        print("Advertencia: Los DataFrames de ingresos o egresos están vacíos")
        
    # Crear conjuntos completos de meses (1-12) para asegurar continuidad
    meses_completos = pd.DataFrame({
        'Año': [int(anio)] * 12,
        'Mes': list(range(1, 13))
    })
    
    # Hacer merge con los datos reales, priorizando tener todos los meses
    df_ingresos_completo = pd.merge(meses_completos, df_ingresos, 
                                   on=['Año', 'Mes'], how='left').infer_objects(copy=False).fillna(0)
    df_egresos_completo = pd.merge(meses_completos, df_egresos, 
                                  on=['Año', 'Mes'], how='left').infer_objects(copy=False).fillna(0)
    
    # Unir ingresos y egresos
    df_resumen = pd.merge(df_ingresos_completo, df_egresos_completo, 
                         on=["Año", "Mes"], how="outer", 
                         suffixes=("_Ingresos", "_Egresos")).fillna(0)
    
    # Calcular balance
    df_resumen["Monto_Ingresos"] = df_resumen["Monto_Ingresos"].apply(lambda x: Decimal(str(x)))
    df_resumen["Monto_Egresos"] = df_resumen["Monto_Egresos"].apply(lambda x: Decimal(str(x)))
    df_resumen["Balance"] = df_resumen["Monto_Ingresos"] - df_resumen["Monto_Egresos"]
    
    # Ordenar por año y mes
    df_resumen = df_resumen.sort_values(by=["Año", "Mes"])
    
    # Convertir los nombres de los meses a texto para mejorar el gráfico
    meses_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    df_resumen['Mes_Nombre'] = df_resumen['Mes'].map(meses_nombres)
    
    # Reordenar columnas
    cols = ['Año', 'Mes', 'Mes_Nombre', 'Monto_Ingresos', 'Monto_Egresos', 'Balance']
    df_resumen = df_resumen[cols]
    
    return df_resumen

# Función para agregar gráficos al Excel con soporte para versiones antiguas de openpyxl
def agregar_graficos(nombre_archivo):
    try:
        wb = openpyxl.load_workbook(nombre_archivo)
        
        # Verificar si la hoja existe
        if "Resumen" not in wb.sheetnames:
            print("La hoja 'Resumen' no existe en el archivo.")
            return
            
        ws = wb["Resumen"]

        # Eliminar gráficos existentes (si los hay)
        for chart in ws._charts:
            ws._charts.remove(chart)

        # Encontrar la última fila con datos
        max_row = ws.max_row
        print(f"Número de filas en la hoja: {max_row}")
        
        # Configurar gráfico de barras
        chart_bar = BarChart()
        chart_bar.title = "Ingresos vs Egresos por Mes"
        chart_bar.y_axis.title = "Monto (ARS)"
        chart_bar.x_axis.title = "Meses"
        chart_bar.style = 10
        
        # Usar nombres de meses como categorías (columna C: Mes_Nombre)
        categories = Reference(ws, min_col=3, min_row=2, max_row=max_row)
        
       # Datos de ingresos (columna D)
        values_ingresos = Reference(ws, min_col=4, max_col=4, min_row=1, max_row=max_row)
        chart_bar.add_data(values_ingresos, titles_from_data=True)

        # Datos de egresos (columna E)
        values_egresos = Reference(ws, min_col=5, max_col=5, min_row=1, max_row=max_row)
        chart_bar.add_data(values_egresos, titles_from_data=True)
        
        chart_bar.set_categories(categories)
        
        try:
            # Intentar configurar marcadores para versiones más recientes
            for s in chart_bar.series:
                s.marker = Marker(symbol="circle", size=8)
        except (AttributeError, TypeError):
            print("Nota: La versión de openpyxl no soporta configuración directa de marcadores.")
            
        # Añadir gráfico de barras a la hoja
        ws.add_chart(chart_bar, "H2")
        
        # Crear gráfico de líneas para el balance
        chart_line = LineChart()
        chart_line.title = "Evolución del Balance"
        chart_line.y_axis.title = "Monto (ARS)"
        chart_line.x_axis.title = "Meses"
        chart_line.style = 13
        
        # Datos del balance (columna F)
        values_balance = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=max_row)
        chart_line.add_data(values_balance, titles_from_data=True)
        chart_line.set_categories(categories)
        
        try:
            # Intentar configurar marcadores para versiones más recientes
            for s in chart_line.series:
                s.marker = Marker(symbol="diamond", size=8)
        except (AttributeError, TypeError):
            print("Nota: La versión de openpyxl no soporta configuración directa de marcadores.")
            
        # Añadir gráfico de líneas a la hoja
        ws.add_chart(chart_line, "H18")
        
        # Guardar el archivo
        wb.save(nombre_archivo)
        print("Gráficos agregados correctamente.")
        
    except Exception as e:
        print(f"Error al agregar gráficos: {e}")

# **MAIN: Ejecutar todo el proceso**
def main():
    # Lista de niveles disponibles
    niveles_disponibles = {
        "1": "Inicial",
        "2": "Primaria",
        "3": "Secundaria",
        "4": "Terciario"
    }
    
    # Mostrar menú de niveles
    print("\nSeleccione el nivel:")
    for key, value in niveles_disponibles.items():
        print(f"{key}. {value}")
    
    # Obtener selección del usuario
    while True:
        nivel_opcion = input("Ingrese el número correspondiente al nivel: ")
        if nivel_opcion in niveles_disponibles:
            nivel = niveles_disponibles[nivel_opcion]
            break
        print("Opción inválida. Por favor, seleccione un número del 1 al 4.")
    
    # Obtener año con valor por defecto (año actual)
    current_year = datetime.now().year  # 2025 based on current date March 03, 2025
    anio_input = input(f"Ingrese el año (presione Enter para usar {current_year}): ")
    anio = anio_input if anio_input else str(current_year)
    
    # Generar nombre del archivo con formato requerido
    current_date = datetime.now().strftime("%d-%m-%Y")  # Format: dd-mm-aaaa
    nombre_archivo = f"{nivel.lower()}-{current_date}-{anio}.xlsx"
    
    niveles = {
        "Inicial": 1,
        "Primaria": 2,
        "Secundaria": 3,
        "Terciario": 4
    }
    
    # Definir consultas (unchanged)
    consultas = {
        "Ingresos": f"SELECT * FROM ingresosinicial WHERE Nivel='{nivel}' AND Año='{anio}' ORDER BY `ingresosinicial`.`Fecha` ASC",
        "Egresos": f"SELECT * FROM egresosinicial WHERE `egresosinicial`.`nivel` = '{nivel}' AND Año='{anio}' ORDER BY `egresosinicial`.`Fecha` ASC",
        "Ingresos Mensuales": f""" 
            SELECT year(`created`) 'Año', 
                month(`created`) 'Mes', 
                SUM(`monto`) 'Monto' 
            FROM `ingresos` 
            WHERE (SELECT alumnos.nivele_id FROM alumnos WHERE alumnos.id = ingresos.alumno_id) = '{niveles[nivel]}' 
                AND YEAR(`created`)='{anio}' 
            GROUP BY year(`created`), month(`created`)
            ORDER BY year(`created`), month(`created`)
        """,
        "Egresos Mensuales": f""" 
            SELECT year(fecha_pago) 'Año', 
                month(fecha_pago) 'Mes', 
                sum(monto_pagado) 'Monto' 
            FROM `erogaciones` 
            WHERE year(fecha_pago)='{anio}' AND nivel='{nivel}' 
            GROUP BY year(fecha_pago), month(fecha_pago)
            ORDER BY year(fecha_pago), month(fecha_pago)
        """,
        "Matricula": f"SELECT * FROM ingresosinicial WHERE Nivel='{nivel}' AND Año='{anio}' AND Concepto LIKE '%Matricula%' ORDER BY `ingresosinicial`.`Fecha` ASC",
        "Nomina de Estudiantes": f"""
            SELECT _nombre_nivel(`nivele_id`) 'Nivel',
                `estado` 'Sala / Grado',
                `apellido` 'Apellido',
                `nombre` 'Nombre',
                `dni` 'DNI / CUIL'
            FROM `alumnos`
            WHERE `nivele_id`='{niveles[nivel]}'
            ORDER BY `estado`, `apellido`, `nombre`
        """
    }

    with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
        dataframes = {}
        hojas_escritas = False  # Bandera para verificar si se escribió algo

        # Ejecutar consultas y guardar en Excel
        for nombre_hoja, query in consultas.items():
            df = ejecutar_consulta(query)
            print(f"Consulta {nombre_hoja}: {len(df)} filas")
            if not df.empty:  # Solo guardar si hay datos
                guardar_en_excel(df, writer, nombre_hoja)
                dataframes[nombre_hoja] = df
                hojas_escritas = True

        # Generar y guardar resumen solo si hay datos
        if "Ingresos Mensuales" in dataframes and "Egresos Mensuales" in dataframes:
            df_resumen = generar_resumen(dataframes["Ingresos Mensuales"], dataframes["Egresos Mensuales"], anio)
            print("Resumen generado con éxito.")
            print(f"Dimensiones del resumen: {df_resumen.shape}")
            df_resumen["Monto_Ingresos"] = pd.to_numeric(df_resumen["Monto_Ingresos"], errors="coerce")
            df_resumen["Monto_Egresos"] = pd.to_numeric(df_resumen["Monto_Egresos"], errors="coerce")
            df_resumen["Balance"] = pd.to_numeric(df_resumen["Balance"], errors="coerce")
            guardar_en_excel(df_resumen, writer, "Resumen")
            hojas_escritas = True

        if not hojas_escritas:
            print("No se generaron datos. Creando hoja vacía para evitar error.")
            pd.DataFrame({"Mensaje": ["No hay datos disponibles"]}).to_excel(writer, sheet_name="Info", index=False)

    # Agregar gráficos solo si hay datos relevantes
    if hojas_escritas:
        agregar_graficos(nombre_archivo)

    print(f"Archivo Excel '{nombre_archivo}' generado con éxito.")

if __name__ == "__main__":
    main()